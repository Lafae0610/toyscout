#!/usr/bin/env python3
"""
泰国潮玩 KOL/KOC 发现工具 - Web版
无需任何API Key，本地直接运行

启动: python3 app.py
访问: http://localhost:5000
"""

import json
import re
import subprocess
import os
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote_plus

from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
DB_FILE = DATA_DIR / "kol_db.json"

DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# ── 数据库 ─────────────────────────────────────────────

def load_db():
    if DB_FILE.exists():
        return json.loads(DB_FILE.read_text(encoding="utf-8"))
    return {"profiles": {}, "last_run": None}


def save_db(db):
    DB_FILE.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")


def init_db():
    """首次运行时导入种子数据"""
    db = load_db()
    if db["profiles"]:
        return db

    for p in SEED_PROFILES:
        key = f"{p['platform'].lower()}:{p['handle'].lower().lstrip('@')}"
        p["discovered_at"] = date.today().isoformat()
        p["status"] = "new"
        db["profiles"][key] = p

    db["last_run"] = datetime.now().isoformat()
    save_db(db)
    return db


# ── 抓取工具 ───────────────────────────────────────────

def curl_fetch(url, extra_headers=None):
    cmd = ["curl", "-sL", "-m", "15",
           "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"]
    if extra_headers:
        for k, v in extra_headers.items():
            cmd += ["-H", f"{k}: {v}"]
    cmd.append(url)
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
        return result.stdout
    except:
        return ""


def fetch_tiktok_oembed(username):
    """通过TikTok oEmbed获取用户最新视频的嵌入代码"""
    # TikTok oEmbed 不需要认证
    profile_url = f"https://www.tiktok.com/@{username.lstrip('@')}"
    url = f"https://www.tiktok.com/oembed?url={quote_plus(profile_url)}"
    html = curl_fetch(url)
    if html:
        try:
            return json.loads(html)
        except:
            pass
    return None


def scrape_profile_meta(url):
    """抓取社交媒体页面的meta信息"""
    html = curl_fetch(url)
    if not html:
        return {}

    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")

    meta = {}
    meta["title"] = soup.title.string if soup.title else ""

    for tag in soup.find_all("meta"):
        prop = tag.get("property", "") or tag.get("name", "")
        content = tag.get("content", "")
        if prop in ("og:description", "description") and content:
            meta["description"] = content
        if prop == "og:image" and content:
            meta["image"] = content
        if prop == "og:title" and content:
            meta["og_title"] = content

    # 提取 JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            ld = json.loads(script.string)
            if isinstance(ld, dict):
                meta["ld_json"] = ld
        except:
            pass

    return meta


# ── 路由 ──────────────────────────────────────────────

@app.route("/")
def index():
    db = load_db()
    profiles = list(db["profiles"].values())

    # 筛选
    platform = request.args.get("platform", "")
    status = request.args.get("status", "")
    search = request.args.get("search", "").lower()
    sort_by = request.args.get("sort", "total")
    filter_type = request.args.get("filter", "")

    if platform:
        profiles = [p for p in profiles if p.get("platform") == platform]
    if status:
        profiles = [p for p in profiles if p.get("status", "new") == status]
    if filter_type == "contactable":
        profiles = [p for p in profiles if p.get("has_business_intent")]
    if search:
        profiles = [p for p in profiles if
                    search in p.get("handle", "").lower() or
                    search in p.get("display_name", "").lower() or
                    search in p.get("bio_summary", "").lower() or
                    search in str(p.get("category_tags", "")).lower()]

    # 排序
    if sort_by == "commercial":
        profiles.sort(key=lambda x: x.get("commercial_score", 0), reverse=True)
    elif sort_by == "conversion":
        profiles.sort(key=lambda x: x.get("conversion_score", 0), reverse=True)
    elif sort_by == "relevance":
        profiles.sort(key=lambda x: x.get("relevance_score", 0), reverse=True)
    else:
        profiles.sort(key=lambda x: (x.get("relevance_score", 0) or 0) + (x.get("commercial_score", 0) or 0) + (x.get("conversion_score", 0) or 0), reverse=True)

    stats = {
        "total": len(db["profiles"]),
        "tiktok": sum(1 for p in db["profiles"].values() if p.get("platform") == "TikTok"),
        "instagram": sum(1 for p in db["profiles"].values() if p.get("platform") == "Instagram"),
        "youtube": sum(1 for p in db["profiles"].values() if p.get("platform") == "YouTube"),
        "contactable": sum(1 for p in db["profiles"].values() if p.get("has_business_intent")),
        "contacted": sum(1 for p in db["profiles"].values() if p.get("status") == "contacted"),
    }

    return render_template("index.html", profiles=profiles, stats=stats,
                         platform=platform, status=status, search=search, sort_by=sort_by)


@app.route("/profile/<path:key>")
def profile_detail(key):
    db = load_db()
    profile = db["profiles"].get(key)
    if not profile:
        return "未找到", 404

    profile["_key"] = key
    return render_template("profile.html", profile=profile)


@app.route("/api/status", methods=["POST"])
def update_status():
    data = request.json
    key = data.get("key")
    new_status = data.get("status")
    note = data.get("note", "")

    db = load_db()
    if key in db["profiles"]:
        db["profiles"][key]["status"] = new_status
        if note:
            db["profiles"][key]["note"] = note
        db["profiles"][key]["updated_at"] = datetime.now().isoformat()
        save_db(db)
        return jsonify({"ok": True})
    return jsonify({"ok": False}), 404


@app.route("/api/scrape", methods=["POST"])
def scrape_profile():
    """手动触发抓取某个profile的页面信息"""
    data = request.json
    key = data.get("key")

    db = load_db()
    profile = db["profiles"].get(key)
    if not profile:
        return jsonify({"ok": False}), 404

    url = profile.get("url", "")
    meta = scrape_profile_meta(url)

    if meta:
        if meta.get("description"):
            profile["scraped_bio"] = meta["description"]
        if meta.get("image"):
            profile["avatar_url"] = meta["image"]
        if meta.get("og_title"):
            profile["scraped_name"] = meta["og_title"]
        profile["last_scraped"] = datetime.now().isoformat()
        save_db(db)

    return jsonify({"ok": True, "meta": meta})


@app.route("/api/add", methods=["POST"])
def add_profile():
    """通过URL手动添加一个新博主"""
    data = request.json
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"ok": False, "error": "请输入URL"}), 400

    # 解析平台和用户名
    platform, handle = "", ""
    if "tiktok.com/@" in url:
        platform = "TikTok"
        m = re.search(r'tiktok\.com/@([^/?#/]+)', url)
        if m:
            handle = f"@{m.group(1)}"
            url = f"https://www.tiktok.com/@{m.group(1)}"
    elif "instagram.com/" in url:
        platform = "Instagram"
        m = re.search(r'instagram\.com/([^/?#]+)', url)
        if m and m.group(1) not in ("p","reel","reels","stories","explore"):
            handle = f"@{m.group(1)}"
            url = f"https://www.instagram.com/{m.group(1)}/"
    elif "youtube.com/" in url:
        platform = "YouTube"
        m = re.search(r'youtube\.com/(?:@|c/|channel/)([^/?#]+)', url)
        if m:
            handle = f"@{m.group(1)}"
            url = f"https://www.youtube.com/@{m.group(1)}"

    if not handle:
        return jsonify({"ok": False, "error": "无法识别，请粘贴 TikTok/Instagram/YouTube 主页链接"}), 400

    key = f"{platform.lower()}:{handle.lower().lstrip('@')}"
    db = load_db()
    if key in db["profiles"]:
        return jsonify({"ok": False, "error": f"{handle} 已存在"}), 400

    # 抓取页面获取基本信息
    meta = scrape_profile_meta(url)
    bio = meta.get("description", "")
    name = meta.get("og_title", "") or handle

    # 用关键词简单判断赛道相关性
    text_all = (bio + " " + name + " " + url).lower()
    toy_keywords = ["toy", "ของเล่น", "blind box", "กล่องสุ่ม", "figure", "ฟิกเกอร์",
                    "art toy", "pop mart", "labubu", "collectible", "unbox", "แกะกล่อง",
                    "molly", "dimoo", "crybaby", "anime", "hobby", "สะสม", "โมเดล"]
    biz_keywords = ["collab", "business", "email", "contact", "ติดต่อ", "work with",
                    "partnership", "dm", "line", "สนใจ", "@gmail", "@hotmail", "agency"]

    matched_tags = []
    tag_map = {"toy": "toy", "ของเล่น": "toy", "blind box": "blind_box", "กล่องสุ่ม": "blind_box",
               "figure": "figure", "ฟิกเกอร์": "figure", "art toy": "art_toy",
               "collectible": "collectible", "unbox": "unboxing", "แกะกล่อง": "unboxing",
               "anime": "anime_merch", "hobby": "hobby", "สะสม": "collectible", "โมเดล": "figure"}
    for kw, tag in tag_map.items():
        if kw in text_all and tag not in matched_tags:
            matched_tags.append(tag)

    relevance = min(100, len([k for k in toy_keywords if k in text_all]) * 15)
    commercial = min(100, len([k for k in biz_keywords if k in text_all]) * 15)
    has_biz = commercial > 20

    # 提取邮箱
    email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', bio)
    email = email_match.group(0) if email_match else ""

    profile = {
        "platform": platform, "handle": handle, "url": url,
        "display_name": name, "followers_est": "待刷新",
        "category_tags": matched_tags or ["toy"],
        "bio_summary": bio[:100] if bio else "已添加，等待刷新获取详细信息",
        "contact_email": email, "contact_line": "", "contact_other": "",
        "has_business_intent": has_biz,
        "business_signals": "已从页面检测到商务信号" if has_biz else "待分析",
        "commercial_score": commercial, "conversion_score": 0,
        "relevance_score": relevance,
        "analysis": "手动添加，可刷新获取更多信息",
        "status": "new", "discovered_at": date.today().isoformat(),
        "source": "manual",
        "avatar_url": meta.get("image", ""),
        "scraped_bio": bio,
    }

    db["profiles"][key] = profile
    save_db(db)
    return jsonify({"ok": True, "key": key, "handle": handle, "platform": platform})


@app.route("/api/refresh", methods=["POST"])
def refresh_all():
    """刷新所有博主的页面信息"""
    db = load_db()
    updated = 0
    total = len(db["profiles"])
    errors = []

    for key, profile in db["profiles"].items():
        url = profile.get("url", "")
        if not url:
            continue

        meta = scrape_profile_meta(url)
        if meta:
            if meta.get("description"):
                profile["scraped_bio"] = meta["description"]
                # 如果原bio是占位文字则更新
                if not profile.get("bio_summary") or profile["bio_summary"].startswith("已添加"):
                    profile["bio_summary"] = meta["description"][:100]
            if meta.get("image"):
                profile["avatar_url"] = meta["image"]
            if meta.get("og_title"):
                profile["scraped_name"] = meta["og_title"]

            # 从bio中提取邮箱
            bio = meta.get("description", "")
            email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', bio)
            if email_match and not profile.get("contact_email"):
                profile["contact_email"] = email_match.group(0)

            # 提取粉丝数（从description中常见格式）
            fans_match = re.search(r'([\d.]+[KMkm]?)\s*(?:Followers|followers|粉丝|ผู้ติดตาม)', bio)
            if fans_match:
                profile["followers_est"] = fans_match.group(1)

            profile["last_scraped"] = datetime.now().isoformat()
            updated += 1
        else:
            errors.append(key)

    db["last_run"] = datetime.now().isoformat()
    save_db(db)
    return jsonify({"ok": True, "updated": updated, "total": total, "errors": len(errors)})


@app.route("/api/refresh/<path:key>", methods=["POST"])
def refresh_one(key):
    """刷新单个博主"""
    db = load_db()
    profile = db["profiles"].get(key)
    if not profile:
        return jsonify({"ok": False}), 404

    url = profile.get("url", "")
    meta = scrape_profile_meta(url)
    if meta:
        if meta.get("description"):
            profile["scraped_bio"] = meta["description"]
            if not profile.get("bio_summary") or profile["bio_summary"].startswith("已添加"):
                profile["bio_summary"] = meta["description"][:100]
        if meta.get("image"):
            profile["avatar_url"] = meta["image"]
        if meta.get("og_title"):
            profile["scraped_name"] = meta["og_title"]
        bio = meta.get("description", "")
        email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', bio)
        if email_match and not profile.get("contact_email"):
            profile["contact_email"] = email_match.group(0)
        profile["last_scraped"] = datetime.now().isoformat()
        save_db(db)
        return jsonify({"ok": True, "meta": meta})

    return jsonify({"ok": True, "meta": {}})


@app.route("/api/delete/<path:key>", methods=["POST"])
def delete_profile(key):
    """删除一个博主"""
    db = load_db()
    if key in db["profiles"]:
        del db["profiles"][key]
        save_db(db)
        return jsonify({"ok": True})
    return jsonify({"ok": False}), 404


@app.route("/api/export")
def export_excel():
    db = load_db()
    status_filter = request.args.get("status", "")

    profiles = list(db["profiles"].values())
    if status_filter:
        profiles = [p for p in profiles if p.get("status") == status_filter]

    wb = Workbook()
    ws = wb.active
    ws.title = "KOL-KOC"

    headers = ["平台", "用户名", "链接", "显示名称", "粉丝量", "赛道标签",
               "Bio摘要", "相关度", "合作意愿", "转化潜力", "邮箱", "LINE",
               "其他联系", "商务信号", "状态", "备注", "发现日期"]

    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, p in enumerate(profiles, 2):
        tags = p.get("category_tags", [])
        if isinstance(tags, list):
            tags = ", ".join(tags)
        row = [p.get("platform",""), p.get("handle",""), p.get("url",""),
               p.get("display_name",""), str(p.get("followers_est","")), tags,
               p.get("bio_summary",""), p.get("relevance_score",0),
               p.get("commercial_score",0), p.get("conversion_score",0),
               p.get("contact_email",""), p.get("contact_line",""),
               p.get("contact_other",""), p.get("business_signals",""),
               p.get("status","new"), p.get("note",""),
               p.get("discovered_at","")]
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)

    ws.freeze_panes = "A2"
    filepath = OUTPUT_DIR / f"thai_kol_export_{date.today().isoformat()}.xlsx"
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)


# ── 种子数据 ──────────────────────────────────────────

SEED_PROFILES = [
    {"platform": "TikTok", "handle": "@witch.blind.box", "url": "https://www.tiktok.com/@witch.blind.box",
     "display_name": "Witch Blind Box", "followers_est": "50K+",
     "category_tags": ["blind_box", "unboxing", "art_toy"],
     "bio_summary": "泰国盲盒开箱达人，专注Pop Mart、Labubu等潮玩开箱",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "频繁品牌合作内容",
     "commercial_score": 75, "conversion_score": 70, "relevance_score": 95,
     "analysis": "盲盒垂类高相关，开箱内容带货力强", "status": "new"},

    {"platform": "TikTok", "handle": "@ppluckytoy", "url": "https://www.tiktok.com/@ppluckytoy",
     "display_name": "PP Lucky Toy", "followers_est": "100K+",
     "category_tags": ["toy", "blind_box", "unboxing"],
     "bio_summary": "泰国玩具开箱博主，覆盖盲盒、手办等品类",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "内容含商品链接和折扣码",
     "commercial_score": 80, "conversion_score": 78, "relevance_score": 90,
     "analysis": "粉丝基数大，带货转化信号明显", "status": "new"},

    {"platform": "TikTok", "handle": "@toytogetherland", "url": "https://www.tiktok.com/@toytogetherland",
     "display_name": "Toy Together Land", "followers_est": "30K+",
     "category_tags": ["toy", "collectible", "unboxing"],
     "bio_summary": "泰国玩具收藏社区账号，测评+开箱",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "有店铺链接，接受品牌合作",
     "commercial_score": 70, "conversion_score": 65, "relevance_score": 88,
     "analysis": "社区型账号，粉丝粘性高", "status": "new"},

    {"platform": "TikTok", "handle": "@tntspace_th", "url": "https://www.tiktok.com/@tntspace_th",
     "display_name": "TNT Space Thailand", "followers_est": "20K+",
     "category_tags": ["art_toy", "figure", "collectible"],
     "bio_summary": "泰国潮玩空间，Art Toy展示和销售",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "店铺账号，有明确商务合作入口",
     "commercial_score": 85, "conversion_score": 72, "relevance_score": 92,
     "analysis": "潮玩零售+内容，商务合作意愿强", "status": "new"},

    {"platform": "TikTok", "handle": "@apple1991s", "url": "https://www.tiktok.com/@apple1991s",
     "display_name": "Apple Toy Review", "followers_est": "15K+",
     "category_tags": ["toy", "unboxing", "blind_box"],
     "bio_summary": "泰国女性玩具博主，盲盒和可爱玩具开箱",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "Bio中标注接合作",
     "commercial_score": 65, "conversion_score": 60, "relevance_score": 85,
     "analysis": "KOC级别，女性受众为主", "status": "new"},

    {"platform": "TikTok", "handle": "@happylin555", "url": "https://www.tiktok.com/@happylin555",
     "display_name": "Happy Lin", "followers_est": "10K+",
     "category_tags": ["blind_box", "unboxing", "hobby"],
     "bio_summary": "盲盒爱好者，日常开箱分享",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "纯爱好者内容",
     "commercial_score": 35, "conversion_score": 45, "relevance_score": 80,
     "analysis": "纯KOC，真实感强但商务化程度低", "status": "new"},

    {"platform": "TikTok", "handle": "@ntmaxx", "url": "https://www.tiktok.com/@ntmaxx",
     "display_name": "NTMaxx", "followers_est": "25K+",
     "category_tags": ["toy", "figure", "anime_merch"],
     "bio_summary": "泰国手办和动漫周边收藏博主",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "有商品推荐和购买链接",
     "commercial_score": 60, "conversion_score": 55, "relevance_score": 82,
     "analysis": "动漫手办垂类，男性受众为主", "status": "new"},

    {"platform": "TikTok", "handle": "@cosmo.vv42", "url": "https://www.tiktok.com/@cosmo.vv42",
     "display_name": "Cosmo VV", "followers_est": "8K+",
     "category_tags": ["art_toy", "blind_box", "collectible"],
     "bio_summary": "潮玩收藏家，设计师玩具爱好者",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "收藏展示为主",
     "commercial_score": 30, "conversion_score": 40, "relevance_score": 88,
     "analysis": "高相关度KOC，适合种草合作", "status": "new"},

    {"platform": "TikTok", "handle": "@jormpavaran", "url": "https://www.tiktok.com/@jormpavaran",
     "display_name": "Jorm Pavaran", "followers_est": "12K+",
     "category_tags": ["blind_box", "unboxing", "art_toy"],
     "bio_summary": "泰国盲盒开箱，Pop Mart系列为主",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "有品牌标记内容",
     "commercial_score": 55, "conversion_score": 50, "relevance_score": 85,
     "analysis": "Pop Mart重度用户，垂类精准", "status": "new"},

    {"platform": "TikTok", "handle": "@december_letter", "url": "https://www.tiktok.com/@december_letter",
     "display_name": "December Letter", "followers_est": "18K+",
     "category_tags": ["blind_box", "art_toy", "gifting"],
     "bio_summary": "盲盒和潮玩礼物推荐博主",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "礼物推荐类内容，天然带货属性",
     "commercial_score": 70, "conversion_score": 68, "relevance_score": 82,
     "analysis": "礼物+潮玩定位，适合节日营销合作", "status": "new"},

    {"platform": "TikTok", "handle": "@ohku.thailand.official", "url": "https://www.tiktok.com/@ohku.thailand.official",
     "display_name": "OHKU Thailand", "followers_est": "5K+",
     "category_tags": ["art_toy", "collectible"],
     "bio_summary": "泰国潮玩品牌/店铺官方账号",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "品牌官方账号，有明确商务入口",
     "commercial_score": 90, "conversion_score": 60, "relevance_score": 95,
     "analysis": "品牌方账号，可作为渠道合作对象", "status": "new"},

    {"platform": "TikTok", "handle": "@yukhuann", "url": "https://www.tiktok.com/@yukhuann",
     "display_name": "Yukhuann", "followers_est": "20K+",
     "category_tags": ["unboxing", "blind_box", "hobby"],
     "bio_summary": "泰国生活方式+盲盒开箱博主",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "生活方式博主，有品牌合作历史",
     "commercial_score": 65, "conversion_score": 62, "relevance_score": 72,
     "analysis": "泛生活方式，潮玩是内容之一", "status": "new"},

    {"platform": "TikTok", "handle": "@groupp.group", "url": "https://www.tiktok.com/@groupp.group",
     "display_name": "Groupp", "followers_est": "10K+",
     "category_tags": ["toy", "unboxing"],
     "bio_summary": "泰国玩具开箱和评测",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "内容创作者，商务化程度待观察",
     "commercial_score": 40, "conversion_score": 45, "relevance_score": 78,
     "analysis": "中小KOC，内容质量稳定", "status": "new"},

    {"platform": "TikTok", "handle": "@u.unsuip", "url": "https://www.tiktok.com/@u.unsuip",
     "display_name": "U Unsuip", "followers_est": "5K+",
     "category_tags": ["blind_box", "unboxing"],
     "bio_summary": "盲盒爱好者，开箱日常",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "纯分享型",
     "commercial_score": 25, "conversion_score": 30, "relevance_score": 75,
     "analysis": "微型KOC，真实感强，适合试用合作", "status": "new"},

    {"platform": "TikTok", "handle": "@thnpoil", "url": "https://www.tiktok.com/@thnpoil",
     "display_name": "Thnpoil", "followers_est": "8K+",
     "category_tags": ["blind_box", "art_toy", "collectible"],
     "bio_summary": "泰国潮玩收藏和盲盒内容",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "收藏展示型",
     "commercial_score": 30, "conversion_score": 35, "relevance_score": 83,
     "analysis": "垂类KOC，收藏向内容", "status": "new"},

    {"platform": "Instagram", "handle": "@taketoys.th", "url": "https://www.instagram.com/taketoys.th/",
     "display_name": "Take Toys Thailand", "followers_est": "15K+",
     "category_tags": ["toy", "art_toy", "collectible"],
     "bio_summary": "泰国玩具店铺，潮玩和收藏品",
     "contact_email": "", "contact_line": "", "contact_other": "DM/Instagram商务功能",
     "has_business_intent": True, "business_signals": "商业账号，有购物功能",
     "commercial_score": 85, "conversion_score": 75, "relevance_score": 90,
     "analysis": "零售型账号，直接带货能力强", "status": "new"},

    {"platform": "Instagram", "handle": "@toystudio_thailand", "url": "https://www.instagram.com/toystudio_thailand/",
     "display_name": "Toy Studio Thailand", "followers_est": "10K+",
     "category_tags": ["art_toy", "figure", "collectible"],
     "bio_summary": "泰国潮玩工作室，设计师玩具展示",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "工作室账号，接受合作咨询",
     "commercial_score": 80, "conversion_score": 65, "relevance_score": 95,
     "analysis": "潮玩行业上游，适合品牌联名", "status": "new"},

    {"platform": "Instagram", "handle": "@collector_thailand", "url": "https://www.instagram.com/collector_thailand/",
     "display_name": "Collector Thailand", "followers_est": "8K+",
     "category_tags": ["collectible", "figure", "toy"],
     "bio_summary": "泰国收藏品社区，手办和潮玩",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "社区型账号，有代购/转售服务",
     "commercial_score": 60, "conversion_score": 55, "relevance_score": 88,
     "analysis": "收藏社区，用户信任度高", "status": "new"},

    {"platform": "Instagram", "handle": "@neverland.collectibles", "url": "https://www.instagram.com/neverland.collectibles/",
     "display_name": "Neverland Collectibles", "followers_est": "12K+",
     "category_tags": ["collectible", "figure", "art_toy"],
     "bio_summary": "泰国收藏品店铺，高端手办和限量潮玩",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "商业账号，有LINE联系方式",
     "commercial_score": 85, "conversion_score": 70, "relevance_score": 92,
     "analysis": "高端潮玩渠道，适合限量款合作", "status": "new"},

    {"platform": "Instagram", "handle": "@thailandtoyexpo", "url": "https://www.instagram.com/thailandtoyexpo/",
     "display_name": "Thailand Toy Expo", "followers_est": "50K+",
     "category_tags": ["toy", "art_toy", "collectible"],
     "bio_summary": "泰国玩具展官方账号，行业活动和资讯",
     "contact_email": "", "contact_line": "", "contact_other": "官网联系表单",
     "has_business_intent": True, "business_signals": "行业展会官方，明确的赞助和合作入口",
     "commercial_score": 95, "conversion_score": 50, "relevance_score": 98,
     "analysis": "行业顶级展会，适合品牌曝光和B端合作", "status": "new"},

    {"platform": "Instagram", "handle": "@toys_encyclopedia_thailand", "url": "https://www.instagram.com/toys_encyclopedia_thailand/",
     "display_name": "Toys Encyclopedia TH", "followers_est": "5K+",
     "category_tags": ["toy", "collectible", "hobby"],
     "bio_summary": "泰国玩具百科，各类玩具资讯和评测",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "资讯型账号",
     "commercial_score": 40, "conversion_score": 35, "relevance_score": 80,
     "analysis": "内容型账号，适合软性植入", "status": "new"},

    {"platform": "Instagram", "handle": "@awwcute.toy", "url": "https://www.instagram.com/awwcute.toy/",
     "display_name": "Aww Cute Toy", "followers_est": "3K+",
     "category_tags": ["toy", "blind_box", "gifting"],
     "bio_summary": "可爱玩具和盲盒分享，礼物推荐",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": False, "business_signals": "小型KOC，纯分享",
     "commercial_score": 25, "conversion_score": 30, "relevance_score": 78,
     "analysis": "微型KOC，适合产品试用和种草", "status": "new"},

    {"platform": "YouTube", "handle": "@toysmania_th", "url": "https://www.youtube.com/@toysmania_th",
     "display_name": "Toys Mania Thailand", "followers_est": "80K+",
     "category_tags": ["toy", "unboxing", "blind_box"],
     "bio_summary": "泰国大型玩具开箱频道，覆盖各类玩具品类",
     "contact_email": "", "contact_line": "", "contact_other": "YouTube商务邮箱",
     "has_business_intent": True, "business_signals": "大量品牌合作视频，有商务邮箱",
     "commercial_score": 85, "conversion_score": 80, "relevance_score": 90,
     "analysis": "头部玩具频道，成熟的商务合作体系", "status": "new"},

    {"platform": "YouTube", "handle": "@reviewtoyth", "url": "https://www.youtube.com/@reviewtoyth",
     "display_name": "Review Toy TH", "followers_est": "30K+",
     "category_tags": ["toy", "unboxing", "figure"],
     "bio_summary": "泰国玩具测评频道，深度评测和对比",
     "contact_email": "", "contact_line": "", "contact_other": "",
     "has_business_intent": True, "business_signals": "测评型内容，有品牌送测历史",
     "commercial_score": 70, "conversion_score": 65, "relevance_score": 88,
     "analysis": "测评型KOL，专业度高，适合新品推广", "status": "new"},
]


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    debug = os.environ.get("FLASK_ENV") != "production"
    print(f"\n  🚀 泰国潮玩KOL/KOC发现工具已启动")
    print(f"  📍 打开浏览器访问: http://localhost:{port}\n")
    app.run(debug=debug, host="0.0.0.0", port=port)
