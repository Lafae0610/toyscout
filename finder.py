#!/usr/bin/env python3
"""
泰国潮玩KOL/KOC自动发现工具
使用 Claude API 智能发现 + Brave Search API 增量搜索
每天跑一次 → 输出Excel

使用方式:
  export ANTHROPIC_API_KEY=sk-ant-xxxxx
  python3 finder.py              # 完整运行
  python3 finder.py --seed-only  # 仅AI种子发现（不需要Brave Key）
"""

import json
import os
import re
import sys
import subprocess
import time
import hashlib
from datetime import datetime, date
from pathlib import Path
from urllib.parse import quote_plus

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

# ── 配置 ──────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
DB_FILE = DATA_DIR / "kol_db.json"

DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

BRAVE_API_KEY = os.environ.get("BRAVE_API_KEY", "")  # 可选，有则用Brave搜索增量发现

# ── 内置种子数据（无需API Key即可使用）─────────────────────
SEED_PROFILES = [
    # ── TikTok 潮玩/盲盒/玩具达人 ──
    {"platform": "TikTok", "handle": "@witch.blind.box", "url": "https://www.tiktok.com/@witch.blind.box",
     "display_name": "Witch Blind Box", "followers_est": "50K+", "category_tags": ["blind_box", "unboxing", "art_toy"],
     "bio_summary": "泰国盲盒开箱达人，专注Pop Mart、Labubu等潮玩开箱", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "频繁品牌合作内容",
     "commercial_score": 75, "conversion_score": 70, "relevance_score": 95, "analysis": "盲盒垂类高相关，开箱内容带货力强"},

    {"platform": "TikTok", "handle": "@ppluckytoy", "url": "https://www.tiktok.com/@ppluckytoy",
     "display_name": "PP Lucky Toy", "followers_est": "100K+", "category_tags": ["toy", "blind_box", "unboxing"],
     "bio_summary": "泰国玩具开箱博主，覆盖盲盒、手办等品类", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "内容含商品链接和折扣码",
     "commercial_score": 80, "conversion_score": 78, "relevance_score": 90, "analysis": "粉丝基数大，带货转化信号明显"},

    {"platform": "TikTok", "handle": "@toytogetherland", "url": "https://www.tiktok.com/@toytogetherland",
     "display_name": "Toy Together Land", "followers_est": "30K+", "category_tags": ["toy", "collectible", "unboxing"],
     "bio_summary": "泰国玩具收藏社区账号，测评+开箱", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "有店铺链接，接受品牌合作",
     "commercial_score": 70, "conversion_score": 65, "relevance_score": 88, "analysis": "社区型账号，粉丝粘性高"},

    {"platform": "TikTok", "handle": "@tntspace_th", "url": "https://www.tiktok.com/@tntspace_th",
     "display_name": "TNT Space Thailand", "followers_est": "20K+", "category_tags": ["art_toy", "figure", "collectible"],
     "bio_summary": "泰国潮玩空间，Art Toy展示和销售", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "店铺账号，有明确商务合作入口",
     "commercial_score": 85, "conversion_score": 72, "relevance_score": 92, "analysis": "潮玩零售+内容，商务合作意愿强"},

    {"platform": "TikTok", "handle": "@apple1991s", "url": "https://www.tiktok.com/@apple1991s",
     "display_name": "Apple Toy Review", "followers_est": "15K+", "category_tags": ["toy", "unboxing", "blind_box"],
     "bio_summary": "泰国女性玩具博主，盲盒和可爱玩具开箱", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "Bio中标注接合作",
     "commercial_score": 65, "conversion_score": 60, "relevance_score": 85, "analysis": "KOC级别，女性受众为主"},

    {"platform": "TikTok", "handle": "@happylin555", "url": "https://www.tiktok.com/@happylin555",
     "display_name": "Happy Lin", "followers_est": "10K+", "category_tags": ["blind_box", "unboxing", "hobby"],
     "bio_summary": "盲盒爱好者，日常开箱分享", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "纯爱好者内容，暂无明显商务信号",
     "commercial_score": 35, "conversion_score": 45, "relevance_score": 80, "analysis": "纯KOC，真实感强但商务化程度低"},

    {"platform": "TikTok", "handle": "@ntmaxx", "url": "https://www.tiktok.com/@ntmaxx",
     "display_name": "NTMaxx", "followers_est": "25K+", "category_tags": ["toy", "figure", "anime_merch"],
     "bio_summary": "泰国手办和动漫周边收藏博主", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "有商品推荐和购买链接",
     "commercial_score": 60, "conversion_score": 55, "relevance_score": 82, "analysis": "动漫手办垂类，男性受众为主"},

    {"platform": "TikTok", "handle": "@cosmo.vv42", "url": "https://www.tiktok.com/@cosmo.vv42",
     "display_name": "Cosmo VV", "followers_est": "8K+", "category_tags": ["art_toy", "blind_box", "collectible"],
     "bio_summary": "潮玩收藏家，设计师玩具爱好者", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "收藏展示为主",
     "commercial_score": 30, "conversion_score": 40, "relevance_score": 88, "analysis": "高相关度KOC，适合种草合作"},

    {"platform": "TikTok", "handle": "@jormpavaran", "url": "https://www.tiktok.com/@jormpavaran",
     "display_name": "Jorm Pavaran", "followers_est": "12K+", "category_tags": ["blind_box", "unboxing", "art_toy"],
     "bio_summary": "泰国盲盒开箱，Pop Mart系列为主", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "有品牌标记内容",
     "commercial_score": 55, "conversion_score": 50, "relevance_score": 85, "analysis": "Pop Mart重度用户，垂类精准"},

    {"platform": "TikTok", "handle": "@december_letter", "url": "https://www.tiktok.com/@december_letter",
     "display_name": "December Letter", "followers_est": "18K+", "category_tags": ["blind_box", "art_toy", "gifting"],
     "bio_summary": "盲盒和潮玩礼物推荐博主", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "礼物推荐类内容，天然带货属性",
     "commercial_score": 70, "conversion_score": 68, "relevance_score": 82, "analysis": "礼物+潮玩定位，适合节日营销合作"},

    {"platform": "TikTok", "handle": "@ohku.thailand.official", "url": "https://www.tiktok.com/@ohku.thailand.official",
     "display_name": "OHKU Thailand", "followers_est": "5K+", "category_tags": ["art_toy", "collectible"],
     "bio_summary": "泰国潮玩品牌/店铺官方账号", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "品牌官方账号，有明确商务入口",
     "commercial_score": 90, "conversion_score": 60, "relevance_score": 95, "analysis": "品牌方账号，可作为渠道合作对象"},

    {"platform": "TikTok", "handle": "@yukhuann", "url": "https://www.tiktok.com/@yukhuann",
     "display_name": "Yukhuann", "followers_est": "20K+", "category_tags": ["unboxing", "blind_box", "hobby"],
     "bio_summary": "泰国生活方式+盲盒开箱博主", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "生活方式博主，有品牌合作历史",
     "commercial_score": 65, "conversion_score": 62, "relevance_score": 72, "analysis": "泛生活方式，潮玩是内容之一"},

    {"platform": "TikTok", "handle": "@groupp.group", "url": "https://www.tiktok.com/@groupp.group",
     "display_name": "Groupp", "followers_est": "10K+", "category_tags": ["toy", "unboxing"],
     "bio_summary": "泰国玩具开箱和评测", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "内容创作者，商务化程度待观察",
     "commercial_score": 40, "conversion_score": 45, "relevance_score": 78, "analysis": "中小KOC，内容质量稳定"},

    {"platform": "TikTok", "handle": "@u.unsuip", "url": "https://www.tiktok.com/@u.unsuip",
     "display_name": "U Unsuip", "followers_est": "5K+", "category_tags": ["blind_box", "unboxing"],
     "bio_summary": "盲盒爱好者，开箱日常", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "纯分享型",
     "commercial_score": 25, "conversion_score": 30, "relevance_score": 75, "analysis": "微型KOC，真实感强，适合试用合作"},

    {"platform": "TikTok", "handle": "@thnpoil", "url": "https://www.tiktok.com/@thnpoil",
     "display_name": "Thnpoil", "followers_est": "8K+", "category_tags": ["blind_box", "art_toy", "collectible"],
     "bio_summary": "泰国潮玩收藏和盲盒内容", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "收藏展示型",
     "commercial_score": 30, "conversion_score": 35, "relevance_score": 83, "analysis": "垂类KOC，收藏向内容"},

    # ── Instagram 潮玩/玩具达人 ──
    {"platform": "Instagram", "handle": "@taketoys.th", "url": "https://www.instagram.com/taketoys.th/",
     "display_name": "Take Toys Thailand", "followers_est": "15K+", "category_tags": ["toy", "art_toy", "collectible"],
     "bio_summary": "泰国玩具店铺，潮玩和收藏品", "contact_email": "", "contact_line": "",
     "contact_other": "DM/Instagram商务功能", "has_business_intent": True, "business_signals": "商业账号，有购物功能",
     "commercial_score": 85, "conversion_score": 75, "relevance_score": 90, "analysis": "零售型账号，直接带货能力强"},

    {"platform": "Instagram", "handle": "@toystudio_thailand", "url": "https://www.instagram.com/toystudio_thailand/",
     "display_name": "Toy Studio Thailand", "followers_est": "10K+", "category_tags": ["art_toy", "figure", "collectible"],
     "bio_summary": "泰国潮玩工作室，设计师玩具展示", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "工作室账号，接受合作咨询",
     "commercial_score": 80, "conversion_score": 65, "relevance_score": 95, "analysis": "潮玩行业上游，适合品牌联名"},

    {"platform": "Instagram", "handle": "@collector_thailand", "url": "https://www.instagram.com/collector_thailand/",
     "display_name": "Collector Thailand", "followers_est": "8K+", "category_tags": ["collectible", "figure", "toy"],
     "bio_summary": "泰国收藏品社区，手办和潮玩", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "社区型账号，有代购/转售服务",
     "commercial_score": 60, "conversion_score": 55, "relevance_score": 88, "analysis": "收藏社区，用户信任度高"},

    {"platform": "Instagram", "handle": "@neverland.collectibles", "url": "https://www.instagram.com/neverland.collectibles/",
     "display_name": "Neverland Collectibles", "followers_est": "12K+", "category_tags": ["collectible", "figure", "art_toy"],
     "bio_summary": "泰国收藏品店铺，高端手办和限量潮玩", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "商业账号，有LINE联系方式",
     "commercial_score": 85, "conversion_score": 70, "relevance_score": 92, "analysis": "高端潮玩渠道，适合限量款合作"},

    {"platform": "Instagram", "handle": "@thailandtoyexpo", "url": "https://www.instagram.com/thailandtoyexpo/",
     "display_name": "Thailand Toy Expo", "followers_est": "50K+", "category_tags": ["toy", "art_toy", "collectible"],
     "bio_summary": "泰国玩具展官方账号，行业活动和资讯", "contact_email": "", "contact_line": "",
     "contact_other": "官网联系表单", "has_business_intent": True, "business_signals": "行业展会官方，明确的赞助和合作入口",
     "commercial_score": 95, "conversion_score": 50, "relevance_score": 98, "analysis": "行业顶级展会，适合品牌曝光和B端合作"},

    {"platform": "Instagram", "handle": "@toys_encyclopedia_thailand", "url": "https://www.instagram.com/toys_encyclopedia_thailand/",
     "display_name": "Toys Encyclopedia TH", "followers_est": "5K+", "category_tags": ["toy", "collectible", "hobby"],
     "bio_summary": "泰国玩具百科，各类玩具资讯和评测", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "资讯型账号",
     "commercial_score": 40, "conversion_score": 35, "relevance_score": 80, "analysis": "内容型账号，适合软性植入"},

    {"platform": "Instagram", "handle": "@awwcute.toy", "url": "https://www.instagram.com/awwcute.toy/",
     "display_name": "Aww Cute Toy", "followers_est": "3K+", "category_tags": ["toy", "blind_box", "gifting"],
     "bio_summary": "可爱玩具和盲盒分享，礼物推荐", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": False, "business_signals": "小型KOC，纯分享",
     "commercial_score": 25, "conversion_score": 30, "relevance_score": 78, "analysis": "微型KOC，适合产品试用和种草"},

    # ── YouTube 玩具/潮玩频道 ──
    {"platform": "YouTube", "handle": "@toysmania_th", "url": "https://www.youtube.com/@toysmania_th",
     "display_name": "Toys Mania Thailand", "followers_est": "80K+", "category_tags": ["toy", "unboxing", "blind_box"],
     "bio_summary": "泰国大型玩具开箱频道，覆盖各类玩具品类", "contact_email": "", "contact_line": "",
     "contact_other": "YouTube商务邮箱", "has_business_intent": True, "business_signals": "大量品牌合作视频，有商务邮箱",
     "commercial_score": 85, "conversion_score": 80, "relevance_score": 90, "analysis": "头部玩具频道，成熟的商务合作体系"},

    {"platform": "YouTube", "handle": "@reviewtoyth", "url": "https://www.youtube.com/@reviewtoyth",
     "display_name": "Review Toy TH", "followers_est": "30K+", "category_tags": ["toy", "unboxing", "figure"],
     "bio_summary": "泰国玩具测评频道，深度评测和对比", "contact_email": "", "contact_line": "",
     "contact_other": "", "has_business_intent": True, "business_signals": "测评型内容，有品牌送测历史",
     "commercial_score": 70, "conversion_score": 65, "relevance_score": 88, "analysis": "测评型KOL，专业度高，适合新品推广"},
]


def load_db() -> dict:
    if DB_FILE.exists():
        return json.loads(DB_FILE.read_text(encoding="utf-8"))
    return {"profiles": {}, "last_run": None, "run_count": 0}


def save_db(db: dict):
    DB_FILE.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")


def profile_key(platform: str, handle: str) -> str:
    return f"{platform.lower()}:{handle.lower().lstrip('@')}"


def curl_fetch(url: str, extra_headers: dict = None) -> str:
    """用系统curl获取网页"""
    cmd = ["curl", "-sL", "-m", "15",
           "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"]
    if extra_headers:
        for k, v in extra_headers.items():
            cmd += ["-H", f"{k}: {v}"]
    cmd.append(url)
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
        return result.stdout
    except:
        return ""


# ── 第一步：Claude AI 种子发现 ──────────────────────────

def discover_with_claude(client: anthropic.Anthropic, existing_handles: set, round_num: int = 1) -> list[dict]:
    """让Claude基于知识库推荐泰国潮玩KOL/KOC"""

    existing_list = ", ".join(list(existing_handles)[:50]) if existing_handles else "无"

    prompt = f"""你是泰国社交媒体行业专家。请帮我发现泰国市场上在"玩具、潮玩(Art Toy)、盲盒(Blind Box)、手办(Figure)、动漫周边、收藏品开箱"赛道有影响力的KOL和KOC。

要求：
1. 覆盖 TikTok、Instagram、YouTube 三个平台
2. 粉丝量从几千到几十万都可以（KOC和KOL都要）
3. 重点关注：有带货转化力的、内容质量好的、有明确商务合作入口的
4. 包括：个人创作者、玩具测评博主、潮玩收藏家、盲盒开箱达人、玩具店铺账号
5. 也包括泰国的 Pop Mart、Labubu、Molly、Dimoo、Crybaby 等潮玩相关内容创作者

已经在库中的账号（请勿重复推荐）：{existing_list}

这是第{round_num}轮发现，请尽量推荐之前没覆盖到的账号。

请输出一个JSON数组，每个元素：
{{
  "platform": "TikTok/Instagram/YouTube",
  "handle": "@xxx（账号用户名）",
  "display_name": "显示名",
  "url": "主页链接",
  "followers_est": "预估粉丝量",
  "category_tags": ["toy", "art_toy", "blind_box", "figure", "collectible", "anime_merch", "unboxing", "hobby"],
  "bio_summary": "一句话介绍（中文）",
  "contact_email": "公开邮箱（如知道）",
  "contact_line": "LINE ID（如知道）",
  "contact_other": "其他联系方式",
  "has_business_intent": true/false,
  "business_signals": "商务合作信号说明",
  "commercial_score": 0-100,
  "conversion_score": 0-100,
  "relevance_score": 0-100,
  "analysis": "简短分析（中文）"
}}

请至少推荐25个账号，越多越好。只输出JSON数组。"""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text
        json_match = re.search(r'\[.*\]', text, re.DOTALL)
        if json_match:
            results = json.loads(json_match.group())
            return [r for r in results if isinstance(r, dict) and r.get("handle")]
    except Exception as e:
        print(f"  ⚠ Claude发现失败: {e}")
    return []


# ── 第二步：Brave Search 增量发现（可选）──────────────────

BRAVE_QUERIES = [
    'thailand tiktok toy unboxing KOL',
    'thailand instagram art toy blind box influencer',
    'thailand youtube toy review figure collectible',
    'ไทย tiktok ของเล่น รีวิว แกะกล่อง',
    'thailand pop mart labubu influencer tiktok',
    'thai toy collector instagram blind box',
    'กล่องสุ่ม ไทย tiktok KOL',
    'ฟิกเกอร์ ไทย youtube review',
]


def brave_search(query: str) -> list[dict]:
    """用Brave Search API搜索（免费2000次/月）"""
    if not BRAVE_API_KEY:
        return []

    url = f"https://api.search.brave.com/res/v1/web/search?q={quote_plus(query)}&count=20&country=th"
    html = curl_fetch(url, {"X-Subscription-Token": BRAVE_API_KEY, "Accept": "application/json"})
    if not html:
        return []

    try:
        data = json.loads(html)
        results = []
        for item in data.get("web", {}).get("results", []):
            href = item.get("url", "")
            info = extract_handle_from_url(href)
            if info:
                info["title"] = item.get("title", "")
                info["description"] = item.get("description", "")
                results.append(info)
        return results
    except:
        return []


def extract_handle_from_url(url: str) -> dict:
    """从URL中提取平台和用户名"""
    info = {"url": url, "platform": "", "handle": ""}

    if "tiktok.com/@" in url:
        info["platform"] = "TikTok"
        match = re.search(r'tiktok\.com/@([^/?#/]+)', url)
        if match:
            info["handle"] = f"@{match.group(1)}"
            info["url"] = f"https://www.tiktok.com/@{match.group(1)}"
    elif "instagram.com/" in url:
        info["platform"] = "Instagram"
        match = re.search(r'instagram\.com/([^/?#]+)', url)
        if match:
            handle = match.group(1)
            if handle not in ("p", "reel", "reels", "stories", "explore", "accounts", "directory", "tv"):
                info["handle"] = f"@{handle}"
                info["url"] = f"https://www.instagram.com/{handle}/"
    elif "youtube.com/" in url:
        info["platform"] = "YouTube"
        match = re.search(r'youtube\.com/(?:@|c/|channel/)([^/?#]+)', url)
        if match:
            info["handle"] = f"@{match.group(1)}"
            info["url"] = f"https://www.youtube.com/@{match.group(1)}"

    return info if info["handle"] else None


# ── 第三步：抓取公开页面补充信息 ─────────────────────────

def scrape_profile_page(url: str) -> str:
    """抓取页面的公开信息文本"""
    try:
        html = curl_fetch(url)
        if not html:
            return ""
        soup = BeautifulSoup(html, "html.parser")

        title = soup.title.string if soup.title else ""
        meta_desc = ""
        for meta in soup.find_all("meta", attrs={"name": "description"}) + soup.find_all("meta", attrs={"property": "og:description"}):
            meta_desc += " " + (meta.get("content", "") or "")

        text_parts = [title, meta_desc.strip()]
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                text_parts.append(script.string or "")
            except:
                pass

        return "\n".join(text_parts)[:2000]
    except:
        return ""


# ── 第四步：Claude 深度分析已发现的账号 ───────────────────

def deep_analyze_with_claude(client: anthropic.Anthropic, profiles: list[dict]) -> list[dict]:
    """让Claude根据抓取到的页面信息深度分析"""

    batch_size = 15
    results = []

    for i in range(0, len(profiles), batch_size):
        batch = profiles[i:i+batch_size]

        profiles_text = ""
        for idx, p in enumerate(batch):
            profiles_text += f"""
--- #{idx+1} {p.get('platform','')} {p.get('handle','')} ---
URL: {p.get('url','')}
已知信息: {p.get('bio_summary','')} | {p.get('analysis','')}
页面抓取: {p.get('page_text','[未抓取]')[:800]}
"""

        prompt = f"""根据以下泰国社交媒体账号的公开页面信息，更新/补充分析。重点提取：联系方式、粉丝数、商务合作信号。

{profiles_text}

对每个账号输出JSON数组，每元素：
{{
  "index": 编号,
  "display_name": "更新后的显示名",
  "followers_est": "从页面提取的粉丝数",
  "bio_summary": "更新后的Bio摘要（中文）",
  "contact_email": "提取到的邮箱",
  "contact_line": "提取到的LINE",
  "contact_other": "其他联系方式（如WhatsApp、Facebook等）",
  "has_business_intent": boolean,
  "business_signals": "更新后的商务信号说明",
  "commercial_score": 0-100,
  "conversion_score": 0-100,
  "relevance_score": 0-100,
  "analysis": "简短分析更新"
}}

只输出JSON数组。"""

        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            )
            text = response.content[0].text
            json_match = re.search(r'\[.*\]', text, re.DOTALL)
            if json_match:
                analyses = json.loads(json_match.group())
                for a in analyses:
                    idx = a.get("index", 1) - 1
                    if 0 <= idx < len(batch):
                        # 合并更新
                        updated = {**batch[idx]}
                        for key in a:
                            if key != "index" and a[key]:
                                updated[key] = a[key]
                        results.append(updated)
                    else:
                        results.append(batch[idx] if idx < len(batch) else a)
            else:
                results.extend(batch)
        except Exception as e:
            print(f"  ⚠ 深度分析失败: {e}")
            results.extend(batch)

        if i + batch_size < len(profiles):
            time.sleep(1)

    return results


# ── 第五步：输出Excel ─────────────────────────────────

def export_excel(profiles: list[dict], filename: str = None):
    """导出结果到Excel"""
    if not filename:
        filename = f"thai_kol_{date.today().isoformat()}.xlsx"

    filepath = OUTPUT_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "KOL-KOC发现"

    headers = [
        "平台", "用户名", "链接", "显示名称", "粉丝量(估)",
        "赛道标签", "Bio摘要", "相关度", "合作意愿", "转化潜力",
        "邮箱", "LINE", "其他联系", "商务信号", "分析",
        "发现日期",
    ]

    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    sorted_profiles = sorted(
        profiles,
        key=lambda x: (x.get("relevance_score", 0) + x.get("commercial_score", 0) + x.get("conversion_score", 0)),
        reverse=True,
    )

    high_fill = PatternFill(start_color="E8F5E9", fill_type="solid")
    mid_fill = PatternFill(start_color="FFF8E1", fill_type="solid")

    for row_idx, p in enumerate(sorted_profiles, 2):
        tags = p.get("category_tags", [])
        if isinstance(tags, list):
            tags = ", ".join(tags)

        row_data = [
            p.get("platform", ""),
            p.get("handle", ""),
            p.get("url", ""),
            p.get("display_name", ""),
            str(p.get("followers_est", "")),
            tags,
            p.get("bio_summary", ""),
            p.get("relevance_score", 0),
            p.get("commercial_score", 0),
            p.get("conversion_score", 0),
            p.get("contact_email", ""),
            p.get("contact_line", ""),
            p.get("contact_other", ""),
            p.get("business_signals", ""),
            p.get("analysis", ""),
            p.get("discovered_at", date.today().isoformat()),
        ]

        total = (p.get("relevance_score", 0) or 0) + (p.get("commercial_score", 0) or 0) + (p.get("conversion_score", 0) or 0)

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if total >= 200:
                cell.fill = high_fill
            elif total >= 120:
                cell.fill = mid_fill

    widths = [10, 24, 42, 18, 12, 30, 35, 8, 8, 8, 28, 15, 22, 35, 45, 12]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64+i) if i <= 26 else "A" + chr(64+i-26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    return filepath


# ── 主流程 ─────────────────────────────────────────────

def run():
    seed_only = "--seed-only" in sys.argv

    print(f"\n{'='*60}")
    print(f"  泰国潮玩 KOL/KOC 自动发现工具")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    client = None
    if api_key and HAS_ANTHROPIC:
        client = anthropic.Anthropic(api_key=api_key)
        print("✅ Claude API 已连接，将启用AI智能发现\n")
    else:
        print("ℹ️  未配置 ANTHROPIC_API_KEY，使用内置种子数据\n")

    db = load_db()
    existing_handles = set(db["profiles"].keys())
    db["run_count"] = db.get("run_count", 0) + 1

    new_profiles = {}

    # ── 内置种子数据导入 ──
    print("📦 第一步：导入种子数据...\n")
    for p in SEED_PROFILES:
        key = profile_key(p.get("platform", ""), p.get("handle", ""))
        if key and key not in existing_handles and key not in new_profiles:
            p_copy = {**p, "discovered_at": date.today().isoformat(), "source": "seed"}
            new_profiles[key] = p_copy
    print(f"  → 种子数据新增 {len(new_profiles)} 个账号")

    # ── AI增量发现（有Key时）──
    if client:
        print(f"\n🤖 第二步：Claude AI 智能发现...\n")
        for round_num in range(1, 3):
            print(f"  第{round_num}轮推荐...")
            discovered = discover_with_claude(client, existing_handles | set(new_profiles.keys()), round_num)
            count_before = len(new_profiles)
            for p in discovered:
                key = profile_key(p.get("platform", ""), p.get("handle", ""))
                if key and key not in existing_handles and key not in new_profiles:
                    p["discovered_at"] = date.today().isoformat()
                    p["source"] = "claude_ai"
                    new_profiles[key] = p
            print(f"  → 本轮推荐 {len(discovered)} 个，新增 {len(new_profiles) - count_before} 个")
            time.sleep(1)

    # ── Brave Search 增量发现（可选）──
    if not seed_only and BRAVE_API_KEY:
        print(f"\n🔍 第三步：Brave Search 增量发现...\n")
        for idx, query in enumerate(BRAVE_QUERIES):
            print(f"  [{idx+1}/{len(BRAVE_QUERIES)}] {query[:50]}...")
            search_results = brave_search(query)
            for info in search_results:
                key = profile_key(info.get("platform", ""), info.get("handle", ""))
                if key and key not in existing_handles and key not in new_profiles:
                    info["discovered_at"] = date.today().isoformat()
                    info["source"] = "brave_search"
                    new_profiles[key] = info
            print(f"       → {len(search_results)} 结果")
            time.sleep(0.5)

    if not new_profiles:
        print("\n⚠ 本次未发现新账号")
        return

    print(f"\n✅ 共发现 {len(new_profiles)} 个新账号\n")

    # ── 抓取公开页面 + AI深度分析（有Key时）──
    if client and not seed_only:
        print("📄 抓取公开页面信息...\n")
        profiles_list = list(new_profiles.values())
        for i, p in enumerate(profiles_list):
            print(f"  [{i+1}/{len(profiles_list)}] {p.get('platform','')} {p.get('handle','')}...")
            p["page_text"] = scrape_profile_page(p.get("url", ""))
            time.sleep(0.5)

        print(f"\n🧠 Claude 深度分析...\n")
        analyzed = deep_analyze_with_claude(client, profiles_list)

        for p in analyzed:
            key = profile_key(p.get("platform", ""), p.get("handle", ""))
            if key in new_profiles:
                p.pop("page_text", None)
                new_profiles[key] = p

    # ── 保存到数据库 ──
    for key, p in new_profiles.items():
        p.pop("page_text", None)
        db["profiles"][key] = p

    db["last_run"] = datetime.now().isoformat()
    save_db(db)

    # ── 输出Excel ──
    print(f"\n📊 生成Excel报告...\n")

    new_list = list(new_profiles.values())
    if new_list:
        path = export_excel(new_list)
        print(f"  ✅ 本次新发现 ({len(new_list)}条): {path}")

    all_data = list(db["profiles"].values())
    if all_data:
        path_all = export_excel(all_data, f"thai_kol_ALL_{date.today().isoformat()}.xlsx")
        print(f"  ✅ 全量汇总 ({len(all_data)}条):   {path_all}")

    print(f"\n{'='*60}")
    print(f"  完成！累计 {len(db['profiles'])} 个KOL/KOC")
    print(f"  结果: {OUTPUT_DIR}")
    if not client:
        print(f"\n  💡 配置 ANTHROPIC_API_KEY 后可解锁：")
        print(f"     - AI智能发现更多KOL（每次50+个）")
        print(f"     - 页面抓取+深度分析")
        print(f"     - 自动提取联系方式")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    run()
